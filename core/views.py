import datetime
from typing import Any, Dict
from django.shortcuts import render, get_object_or_404
from django.http import HttpRequest, HttpResponse
from django.urls import reverse
import openpyxl
from django.db import transaction
from .models import Kit
from django.views.generic import ListView, DetailView
from django.db.models import Q
from .forms import KitFilterForm
from django.core.paginator import Paginator


def home_page_view(request):

    # If this is a POST request then process the Form data
    if request.method == 'POST':
        file = request.FILES['archive']

        wb = openpyxl.load_workbook(file)

        # getting a particular sheet by name out of many sheets
        sheets = wb.sheetnames

        for sheet in sheets:
            worksheet = wb[sheet]

            for index, row in enumerate(worksheet.iter_rows()):
                if index == 0:
                    continue

                new_kit = Kit.objects.create(
                    type_kit=sheet,
                    identification=row[0].value,
                    code=row[1].value,
                    price=row[2].value,
                    roof=row[3].value,
                    connection=row[4].value,

                    quantity_modules=row[5].value,
                    module_model=row[6].value,
                    module_unit_Wp_power=row[7].value,

                    max_overload=row[8].value,
                    kWp=row[9].value,

                    inverter_1_quantity=row[10].value,
                    inverter_1_model=row[11].value,
                    inverter_2_quantity=row[12].value,
                    inverter_2_model=row[13].value,

                    amount_of_red_cables=row[14].value,
                    model_red_cables=row[15].value,
                    amount_of_black_cables=row[16].value,
                    model_black_cables=row[17].value,

                    quantity_pairs_connector=row[18].value,
                    connector_pair_model=row[19].value,

                    quantity_stringbox=row[20].value,
                    model_stringbox=row[21].value,

                    quantity_structure_1=row[22].value,
                    model_structure_1=row[23].value,

                    quantity_structure_2=row[24].value,
                    model_structure_2=row[25].value,

                    quantity_structure_3=row[26].value,
                    model_structure_3=row[27].value,

                    quantity_structure_4=row[28].value,
                    model_structure_4=row[29].value,

                    quantity_structure_5=row[30].value,
                    model_structure_5=row[31].value,

                    inverter_brand=row[32].value,
                    module_brand=row[33].value,
                )

            new_kit.save()

    context = {
        'form': '',
    }

    return render(request, 'home.html', context)


def filter_kits_view(request):

    if request.method == 'GET':
        form = KitFilterForm(request.GET)

        type_kit = request.GET.get('type_kit')
        inverter_brand = request.GET.get('inverter_brand')
        module_brand = request.GET.get('module_brand')
        code = request.GET.get('code')
        amount_of_red_cables = request.GET.get('amount_of_red_cables')
        amount_of_black_cables = request.GET.get('amount_of_black_cables')
        quantity_stringbox = request.GET.get('quantity_stringbox')
        roof = request.GET.get('roof')

        kits = Kit.objects.all()

        if type_kit:
            kits = kits.filter(Q(type_kit=type_kit)) 

        if inverter_brand:
            kits = kits.filter(Q(inverter_brand=inverter_brand))

        if module_brand:
            kits = kits.filter(Q(module_brand=module_brand))

        if roof:
            kits = kits.filter(Q(roof=roof))
        
        if code:
            kits = kits.filter(Q(code__icontains=code))

        if amount_of_red_cables:
            kits = kits.filter(Q(amount_of_red_cables=amount_of_red_cables))

        if amount_of_black_cables:
            kits = kits.filter(Q(amount_of_black_cables=amount_of_black_cables))

        if quantity_stringbox:
            kits = kits.filter(Q(quantity_stringbox=quantity_stringbox))

        context = {
            'object_list': kits,
            'form': form,
        }

    return render(request, 'kit-filters.html', context)


class KitListView(ListView):
    model = Kit
    queryset = Kit.objects.all()
    allow_empty = True
    template_name = 'list-kits.html'
    paginate_by = 15


class KitDetailsView(DetailView):
    model = Kit
    allow_empty = True
    template_name = 'kit-details.html'


class KitSearchView(ListView):
    model = Kit
    template_name = 'search.html'
    queryset = Kit.objects.all()
    paginate_by = 15

    def get_queryset(self, *args, **kwargs):
        search_term = self.request.GET.get('q', '')
        qs = super().get_queryset(*args, **kwargs)
        qs = qs.filter(
            Q(
                Q(identification__icontains=search_term),
            )
        )
        return qs

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        search_term = self.request.GET.get('q', '')

        context.update({
            'search_term': search_term,
        })

        return context
